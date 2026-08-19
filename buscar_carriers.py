"""
FMCSA Carrier Lead Generator
Busca carriers nuevos con 1-3 camiones (semi trucks)
Datos oficiales del Departamento de Transporte de EE.UU.
"""

import os
import sys
from datetime import datetime, timedelta

try:
    import pandas as pd
    import requests
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando librerias necesarias...")
    os.system(f"{sys.executable} -m pip install pandas requests openpyxl -q")
    import pandas as pd
    import requests
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

# ── CONFIGURACION ─────────────────────────────────────────────────────────────

MESES_NUEVOS   = 12      # Carriers registrados en los ultimos N meses
MIN_CAMIONES   = 1
MAX_CAMIONES   = 3
SOLO_DRY_VAN   = False   # True = solo general freight; False = todos los carriers
ARCHIVO_SALIDA = "leads_carriers.xlsx"

API_BASE = "https://data.transportation.gov/resource/az4n-8mr2.csv"
LIMIT    = 50000         # Registros por peticion (maximo del API)

# ── CALCULAR FECHA LIMITE ─────────────────────────────────────────────────────

fecha_limite = (datetime.now() - timedelta(days=MESES_NUEVOS * 30)).strftime("%Y%m%d")
print(f"\nBuscando carriers registrados desde: {fecha_limite[:4]}-{fecha_limite[4:6]}-{fecha_limite[6:]}")

# ── QUERY AL API ──────────────────────────────────────────────────────────────

COLUMNAS = (
    "dot_number,docket1,legal_name,phone,phy_street,phy_city,phy_state,phy_zip,"
    "power_units,add_date,interstate_beyond_100_miles,"
    "docket1prefix,docket1_status_code,"
    "crgo_genfreight,crgo_metalsheet,crgo_bldgmat,crgo_drybulk,crgo_construct"
)

where = (
    f"add_date >= '{fecha_limite}' "
    f"AND status_code = 'A' "
    f"AND docket1prefix = 'MC' "
    f"AND docket1_status_code = 'A' "
    f"AND interstate_beyond_100_miles != '0' "
    f"AND power_units::number >= {MIN_CAMIONES} "
    f"AND power_units::number <= {MAX_CAMIONES}"
)

if SOLO_DRY_VAN:
    where += " AND crgo_genfreight = 'X'"

print("Descargando datos de FMCSA (puede tardar 1-2 minutos)...\n")

todos = []
offset = 0

while True:
    try:
        r = requests.get(
            API_BASE,
            params={
                "$select": COLUMNAS,
                "$where":  where,
                "$order":  "add_date DESC",
                "$limit":  str(LIMIT),
                "$offset": str(offset),
            },
            timeout=60,
        )
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"Error de conexion: {e}")
        sys.exit(1)

    from io import StringIO
    chunk = pd.read_csv(StringIO(r.text), dtype=str)

    if chunk.empty:
        break

    todos.append(chunk)
    offset += len(chunk)
    print(f"  Descargados: {offset:,} carriers...")

    if len(chunk) < LIMIT:
        break

if not todos:
    print("\nNo se encontraron carriers con esos criterios.")
    sys.exit(0)

df = pd.concat(todos, ignore_index=True)
print(f"\nTotal descargado: {len(df):,} carriers")

# Filtrar solo los que tienen MC number
df["docket1"] = df["docket1"].fillna("").astype(str).str.strip()
df = df[df["docket1"] != ""].copy()
print(f"Con MC number:    {len(df):,} carriers")

# ── LIMPIAR Y PREPARAR ────────────────────────────────────────────────────────

# Tipo de carga
def tipo_carga(row):
    tipos = []
    if row.get("crgo_genfreight") == "X":
        tipos.append("General Freight / Dry Van")
    if row.get("crgo_metalsheet") == "X":
        tipos.append("Metal / Acero")
    if row.get("crgo_bldgmat") == "X":
        tipos.append("Materiales de Construccion")
    if row.get("crgo_construct") == "X":
        tipos.append("Construccion")
    if row.get("crgo_drybulk") == "X":
        tipos.append("Dry Bulk")
    return ", ".join(tipos) if tipos else "Otro"

df["Tipo de Carga"] = df.apply(tipo_carga, axis=1)

# Formatear telefono
def fmt_phone(p):
    p = str(p).strip() if pd.notna(p) else ""
    digits = "".join(c for c in p if c.isdigit())
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return p or "No disponible"

df["phone"] = df["phone"].apply(fmt_phone)

# Formatear fecha
def fmt_date(d):
    d = str(d).strip() if pd.notna(d) else ""
    if len(d) == 8 and d.isdigit():
        return f"{d[4:6]}/{d[6:]}/{d[:4]}"
    return d

df["add_date"] = df["add_date"].apply(fmt_date)

# Direccion completa
df["Direccion"] = (
    df["phy_street"].fillna("") + ", " +
    df["phy_city"].fillna("") + ", " +
    df["phy_state"].fillna("") + " " +
    df["phy_zip"].fillna("")
).str.strip(", ")

# ── ORDENAR: DRY VAN PRIMERO ─────────────────────────────────────────────────

df["_prioridad"] = df["crgo_genfreight"].apply(lambda x: 0 if x == "X" else 1)
df = df.sort_values(["_prioridad", "add_date"], ascending=[True, False])

# ── COLUMNAS FINALES ──────────────────────────────────────────────────────────

MESSAGE = (
    "Hi! I found your company on the FMCSA database and wanted to reach out. "
    "My name is Fernando Pabon, I'm a bilingual (English/Spanish) truck dispatcher. "
    "I'm currently looking for 2 Dry Van or Flatbed owner operators — I offer the first 5 loads "
    "completely free, no commitment. If this sounds interesting, I'd love to schedule a quick call. "
    "If not, no worries, I won't bother you again. Have a great day!"
)

def make_sms_link(phone):
    digits = "".join(c for c in str(phone) if c.isdigit())
    if len(digits) < 10:
        return ""
    number = digits[-10:]
    return f"tel:{number}"

df["_sms_link"] = df["phone"].apply(make_sms_link)

df["MC Number"] = "MC-" + df["docket1"]

resultado = df[[
    "MC Number", "legal_name", "phone",
    "_sms_link", "Direccion", "phy_state",
    "power_units", "add_date",
    "Tipo de Carga"
]].copy()

resultado.columns = [
    "MC Number", "Nombre de la Empresa", "Telefono",
    "Enviar Mensaje", "Direccion", "Estado",
    "Num Camiones", "Fecha Registro",
    "Tipo de Carga"
]

# ── EXPORTAR A EXCEL ──────────────────────────────────────────────────────────

print(f"\nExportando {len(resultado):,} leads a Excel...")

with pd.ExcelWriter(ARCHIVO_SALIDA, engine="openpyxl") as writer:
    resultado.to_excel(writer, index=False, sheet_name="Leads")
    ws = writer.sheets["Leads"]

    # Formato de encabezado
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Columna "Enviar Mensaje" (col D = index 3) como hiperlink
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        link_cell = row[3]
        url = str(link_cell.value or "")
        if url.startswith("sms:"):
            link_cell.hyperlink = url
            link_cell.value = "Tap to Text"
            link_cell.font = Font(color="0563C1", underline="single", size=11)

    # Colorear filas dry van en verde claro
    green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cargo_cell = row[8]  # columna Tipo de Carga (ahora col 9)
        if "Dry Van" in str(cargo_cell.value):
            for cell in row:
                if not cell.font or not cell.font.underline:
                    cell.fill = green

    # Ancho de columnas
    anchos = [12, 35, 18, 15, 40, 8, 13, 15, 30]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Congelar encabezado
    ws.freeze_panes = "A2"

ruta = os.path.abspath(ARCHIVO_SALIDA)
dry_van_count = len(resultado[resultado["Tipo de Carga"].str.contains("Dry Van", na=False)])
sms_count = len(resultado[resultado["Enviar Mensaje"] != ""])

print(f"\n{'='*55}")
print(f"  LISTO!")
print(f"{'='*55}")
print(f"  Total leads:           {len(resultado):,}")
print(f"  Dry Van / Gen Freight: {dry_van_count:,}  (filas verdes)")
print(f"  Con link para textar:  {sms_count:,}")
print(f"  Archivo:               {ruta}")
print(f"{'='*55}\n")

os.startfile(ruta)
